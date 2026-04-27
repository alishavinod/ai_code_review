"""
Generate manual audit Excel file with 3 sheets (one per pipeline stage).
Uses the SAME 20 examples across all 3 sheets for consistency.
Includes both old_code and new_code for full diff context.

Output: data/audit/manual_audit.xlsx
"""

import json
import random
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAW_DATASET_FILE      = Path("data/raw/codereviewqa_sample_300.json")
AI_SUGGESTIONS_FILE   = Path("data/generated/ai_suggestions.jsonl")
SEMANTIC_MATCHES_FILE = Path("data/generated/semantic_matches.jsonl")
NOISE_LABELS_FILE     = Path("data/generated/noise_labels.jsonl")
OUTPUT_FILE           = Path("data/audit/manual_audit.xlsx")

SAMPLES_PER_STAGE = 20
random.seed(42)

HEADER_BG   = "1F4E79"
FILL_YELLOW = "FFF2CC"
FILL_GRAY   = "F2F2F2"

LABEL_COLORS = {
    "valid":           "E2EFDA",
    "unmatched":       "FCE4D6",
    "uncertain":       "FDEBD0",
    "trivial":         "DAEEF3",
    "incorrect":       "FCE4D6",
    "context-missing": "FDEBD0",
    "irrelevant":      "F2F2F2",
}

def load_jsonl(path: Path) -> list:
    records = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records

def load_raw(path: Path) -> dict:
    """Load raw dataset and index by position (0-based)."""
    records = {}
    with open(path, encoding="utf-8") as f:
        for i, line in enumerate(f):
            line = line.strip()
            if line:
                records[i] = json.loads(line)
    return records

def header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def data_style(cell, bg="FFFFFF", bold=False):
    cell.font = Font(name="Arial", size=9, bold=bold)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(vertical="top", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def pick_samples(noise_labels):
    noise_examples = [d for d in noise_labels if d["noise_type"] != "valid"]
    by_label = defaultdict(list)
    for d in noise_examples:
        by_label[d["noise_type"]].append(d)

    sample = []
    per_label = SAMPLES_PER_STAGE // len(by_label)
    for label, examples in by_label.items():
        sample.extend(random.sample(examples, min(per_label, len(examples))))
    sample = sample[:SAMPLES_PER_STAGE]
    random.shuffle(sample)
    return sample

def add_sheet_stage1(wb, sample, ai_by_index, raw_by_index):
    ws = wb.create_sheet("Stage1 - AI Quality")

    headers = ["#", "Index", "Lang", "Code BEFORE (old)",
               "Code AFTER (new)", "AI Suggestion",
               "Your Label\n(good/off-target/incoherent)", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)
    ws.row_dimensions[1].height = 35

    for i, ex in enumerate(sample):
        row = i + 2
        bg = FILL_GRAY if i % 2 == 0 else "FFFFFF"
        ai = ai_by_index.get(ex["index"], {})
        raw = raw_by_index.get(ex["index"], {})
        values = [
            i + 1,
            ex["index"],
            ex.get("language", "?"),
            raw.get("old", ex.get("old_code", ""))[:600],
            raw.get("new", "")[:600],
            ai.get("ai_suggestion", ex.get("ai_suggestion", "")),
            "",
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            data_style(cell, bg=FILL_YELLOW if col == 7 else bg)
            cell.border = thin_border()
        ws.row_dimensions[row].height = 100

    for col, w in enumerate([4, 7, 6, 45, 45, 55, 22, 25], 1):
        set_col_width(ws, col, w)
    ws.freeze_panes = "A2"

def add_sheet_stage2(wb, sample, sem_by_index, raw_by_index):
    ws = wb.create_sheet("Stage2 - Semantic Match")

    headers = ["#", "Index", "Lang", "Code BEFORE (old)", "Code AFTER (new)",
               "Human Review (ground truth)", "AI Suggestion",
               "Assigned Label", "Your Label\n(valid/unmatched/uncertain)",
               "Agree?\n(y/n)", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)
    ws.row_dimensions[1].height = 35

    for i, ex in enumerate(sample):
        row = i + 2
        sem = sem_by_index.get(ex["index"], {})
        raw = raw_by_index.get(ex["index"], {})
        assigned = sem.get("match_label", "unmatched")
        bg = LABEL_COLORS.get(assigned, "FFFFFF")
        values = [
            i + 1,
            ex["index"],
            ex.get("language", "?"),
            raw.get("old", ex.get("old_code", ""))[:600],
            raw.get("new", "")[:600],
            sem.get("human_review", ex.get("human_review", "")),
            sem.get("ai_suggestion", ex.get("ai_suggestion", "")),
            assigned,
            "",
            "",
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col == 9:
                data_style(cell, bg=FILL_YELLOW)
            elif col == 8:
                data_style(cell, bg=bg, bold=True)
            else:
                data_style(cell, bg="FFFFFF" if i % 2 == 0 else FILL_GRAY)
            cell.border = thin_border()
        ws.row_dimensions[row].height = 100

    for col, w in enumerate([4, 7, 6, 45, 45, 40, 55, 14, 22, 10, 25], 1):
        set_col_width(ws, col, w)
    ws.freeze_panes = "A2"

def add_sheet_stage3(wb, sample, raw_by_index):
    ws = wb.create_sheet("Stage3 - Noise Label")

    headers = ["#", "Index", "Lang", "Code BEFORE (old)", "Code AFTER (new)",
               "AI Suggestion", "Human Review (ground truth)",
               "Assigned Label",
               "Your Label\n(trivial/incorrect/\ncontext-missing/irrelevant)",
               "Agree?\n(y/n)", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)
    ws.row_dimensions[1].height = 40

    for i, ex in enumerate(sample):
        row = i + 2
        assigned = ex["noise_type"]
        raw = raw_by_index.get(ex["index"], {})
        bg = LABEL_COLORS.get(assigned, "FFFFFF")
        values = [
            i + 1,
            ex["index"],
            ex.get("language", "?"),
            raw.get("old", ex.get("old_code", ""))[:600],
            raw.get("new", "")[:600],
            ex.get("ai_suggestion", ""),
            ex.get("human_review", ""),
            assigned,
            "",
            "",
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col == 9:
                data_style(cell, bg=FILL_YELLOW)
            elif col == 8:
                data_style(cell, bg=bg, bold=True)
            else:
                data_style(cell, bg="FFFFFF" if i % 2 == 0 else FILL_GRAY)
            cell.border = thin_border()
        ws.row_dimensions[row].height = 110

    for col, w in enumerate([4, 7, 6, 45, 45, 55, 40, 16, 22, 10, 25], 1):
        set_col_width(ws, col, w)
    ws.freeze_panes = "A2"

def main():
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    print("Loading data...")
    ai_suggestions   = load_jsonl(AI_SUGGESTIONS_FILE)
    semantic_matches = load_jsonl(SEMANTIC_MATCHES_FILE)
    noise_labels     = load_jsonl(NOISE_LABELS_FILE)
    raw_by_index     = load_raw(RAW_DATASET_FILE)

    ai_by_index  = {d["index"]: d for d in ai_suggestions}
    sem_by_index = {d["index"]: d for d in semantic_matches}

    sample = pick_samples(noise_labels)
    indices = [ex["index"] for ex in sample]
    print(f"Sampled {len(sample)} examples: indices {indices}")

    wb = Workbook()
    wb.remove(wb.active)

    print("Building sheets...")
    add_sheet_stage1(wb, sample, ai_by_index, raw_by_index)
    add_sheet_stage2(wb, sample, sem_by_index, raw_by_index)
    add_sheet_stage3(wb, sample, raw_by_index)

    wb.save(OUTPUT_FILE)
    print(f"\nDone. Saved to {OUTPUT_FILE}")
    print("All 3 sheets use the same 20 examples with full before/after diff.")
    print("Fill in the yellow columns in each sheet.")

if __name__ == "__main__":
    main()